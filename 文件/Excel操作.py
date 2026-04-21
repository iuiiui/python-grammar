# ==================== Excel操作常用方法 ====================
# 需要安装: pip install openpyxl

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== 1. 创建Excel文件 ====================
print("1. 创建Excel文件")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "工作表1"

# 写入数据
ws['A1'] = '姓名'
ws['B1'] = '年龄'
ws['C1'] = '城市'

# 使用 append 添加行数据
ws.append(['张三', 25, '北京'])
ws.append(['李四', 30, '上海'])
ws.append(['王五', 28, '广州'])

# 保存文件
wb.save('示例.xlsx')
print("Excel文件创建成功\n")

# ==================== 2. 读取Excel文件 ====================
print("2. 读取Excel文件")
wb_read = openpyxl.load_workbook('示例.xlsx')
ws_read = wb_read.active

# 读取单个单元格
print(f"A1单元格: {ws_read['A1'].value}")
print(f"B2单元格: {ws_read.cell(row=2, column=2).value}")

# 遍历所有行
print("\n所有数据:")
for row in ws_read.iter_rows(min_row=1, values_only=True):
    print(row)

# 遍历指定范围
print("\n指定范围数据 (A1:C2):")
for row in ws_read.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print(row)

# 获取最大行数和列数
print(f"\n最大行数: {ws_read.max_row}")
print(f"最大列数: {ws_read.max_column}")
print()

# ==================== 3. 样式设置 ====================
print("3. 样式设置")
wb_style = openpyxl.Workbook()
ws_style = wb_style.active
ws_style.title = "样式示例"

# 写入数据
ws_style['A1'] = '标题'
ws_style['A2'] = '内容1'
ws_style['B2'] = '内容2'

# 设置字体
bold_font = Font(name='微软雅黑', size=14, bold=True, color='FF0000')
ws_style['A1'].font = bold_font

normal_font = Font(name='宋体', size=11)
ws_style['A2'].font = normal_font
ws_style['B2'].font = normal_font

# 设置背景填充
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws_style['A1'].fill = yellow_fill

# 设置对齐方式
center_alignment = Alignment(horizontal='center', vertical='center')
ws_style['A1'].alignment = center_alignment
ws_style['A2'].alignment = center_alignment

# 设置边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ws_style['A1'].border = thin_border
ws_style['A2'].border = thin_border
ws_style['B2'].border = thin_border

# 设置行高和列宽
ws_style.row_dimensions[1].height = 30
ws_style.column_dimensions['A'].width = 15
ws_style.column_dimensions['B'].width = 15

wb_style.save('样式示例.xlsx')
print("样式设置完成\n")

# ==================== 4. 多个工作表操作 ====================
print("4. 多个工作表操作")
wb_multi = openpyxl.Workbook()

# 创建第一个工作表
ws1 = wb_multi.active
ws1.title = "学生信息"
ws1.append(['姓名', '成绩'])
ws1.append(['张三', 90])
ws1.append(['李四', 85])

# 创建第二个工作表
ws2 = wb_multi.create_sheet("成绩统计")
ws2['A1'] = '科目'
ws2['B1'] = '平均分'
ws2.append(['数学', 87.5])
ws2.append(['英语', 88.0])

# 创建第三个工作表（指定位置）
ws3 = wb_multi.create_sheet("备注", 0)  # 插入到最前面
ws3['A1'] = '说明'
ws3['B1'] = '这是备注信息'

# 查看所有工作表名称
print("工作表列表:", wb_multi.sheetnames)

# 通过名称获取工作表
ws_by_name = wb_multi['学生信息']
print(f"学生信息表 A1: {ws_by_name['A1'].value}")

wb_multi.save('多工作表示例.xlsx')
print("多工作表创建完成\n")

# ==================== 5. 公式和函数 ====================
print("5. 公式和函数")
wb_formula = openpyxl.Workbook()
ws_formula = wb_formula.active
ws_formula.title = "公式示例"

# 写入数据
ws_formula['A1'] = '数值1'
ws_formula['B1'] = '数值2'
ws_formula['C1'] = '结果'
ws_formula.append([10, 20])
ws_formula.append([30, 40])
ws_formula.append([50, 60])

# 添加公式
ws_formula['C2'] = '=A2+B2'  # 求和
ws_formula['C3'] = '=A3*B3'  # 乘积
ws_formula['C4'] = '=AVERAGE(A2:A4)'  # 平均值

# 在底部添加汇总
ws_formula['A6'] = '总计'
ws_formula['B6'] = '=SUM(B2:B4)'

wb_formula.save('公式示例.xlsx')
print("公式添加完成\n")

# ==================== 6. 合并单元格 ====================
print("6. 合并单元格")
wb_merge = openpyxl.Workbook()
ws_merge = wb_merge.active
ws_merge.title = "合并示例"

# 合并单元格
ws_merge.merge_cells('A1:C1')
ws_merge['A1'] = '合并的标题'
ws_merge['A1'].alignment = Alignment(horizontal='center', vertical='center')

# 合并区域
ws_merge.merge_cells('A2:B3')
ws_merge['A2'] = '合并区域'

wb_merge.save('合并单元格示例.xlsx')
print("合并单元格完成\n")

# ==================== 7. 插入和删除行列 ====================
print("7. 插入和删除行列")
wb_insert = openpyxl.Workbook()
ws_insert = wb_insert.active

# 初始数据
ws_insert.append(['A1', 'B1', 'C1'])
ws_insert.append(['A2', 'B2', 'C2'])
ws_insert.append(['A3', 'B3', 'C3'])

# 在第2行前插入一行
ws_insert.insert_rows(2)
ws_insert['A2'] = '插入的行'

# 在第2列前插入一列
ws_insert.insert_cols(2)
ws_insert['B1'] = '插入的列'

# 删除第3行
# ws_insert.delete_rows(3)

# 删除第2列
# ws_insert.delete_cols(2)

wb_insert.save('插入删除示例.xlsx')
print("插入行列完成\n")

# ==================== 8. 图表制作 ====================
print("8. 图表制作")
wb_chart = openpyxl.Workbook()
ws_chart = wb_chart.active
ws_chart.title = "图表数据"

# 准备数据
ws_chart['A1'] = '月份'
ws_chart['B1'] = '销售额'
data = [
    ['1月', 1000],
    ['2月', 1500],
    ['3月', 1200],
    ['4月', 1800],
    ['5月', 2000],
]
for row in data:
    ws_chart.append(row)

# 创建柱状图
from openpyxl.chart import BarChart, Reference

chart = BarChart()
chart.title = "月度销售额"
chart.x_axis.title = "月份"
chart.y_axis.title = "销售额"

# 设置数据范围
data_ref = Reference(ws_chart, min_col=2, min_row=1, max_row=6, max_col=2)
cat_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=6)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)

# 将图表添加到工作表
ws_chart.add_chart(chart, "D2")

wb_chart.save('图表示例.xlsx')
print("图表创建完成\n")

# ==================== 9. 数据验证 ====================
print("9. 数据验证")
wb_validation = openpyxl.Workbook()
ws_validation = wb_validation.active
ws_validation.title = "数据验证"

from openpyxl.worksheet.datavalidation import DataValidation

# 创建下拉列表验证
dv = DataValidation(type="list", formula1='"选项1,选项2,选项3"')
dv.error = "请输入有效选项"
dv.errorTitle = "输入错误"
dv.prompt = "请从下拉列表中选择"
dv.promptTitle = "提示"

# 应用到单元格
ws_validation.add_data_validation(dv)
dv.add('A1:A10')

ws_validation['A1'] = '请选择'

wb_validation.save('数据验证示例.xlsx')
print("数据验证设置完成\n")

# ==================== 10. 批量数据处理 ====================
print("10. 批量数据处理")
wb_batch = openpyxl.Workbook()
ws_batch = wb_batch.active
ws_batch.title = "批量数据"

# 批量写入数据
headers = ['ID', '姓名', '分数', '等级']
ws_batch.append(headers)

students = [
    (1, '张三', 95),
    (2, '李四', 82),
    (3, '王五', 67),
    (4, '赵六', 91),
    (5, '孙七', 78),
]

for student in students:
    # 计算等级
    if student[2] >= 90:
        grade = '优秀'
    elif student[2] >= 80:
        grade = '良好'
    elif student[2] >= 70:
        grade = '中等'
    else:
        grade = '及格'
    
    ws_batch.append([student[0], student[1], student[2], grade])

# 为表头添加样式
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

for cell in ws_batch[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

wb_batch.save('批量处理示例.xlsx')
print("批量数据处理完成\n")

# ==================== 实用工具函数 ====================
def read_excel_to_list(filepath, sheet_name=None):
    """读取Excel文件为列表"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return data

def write_list_to_excel(data, filepath, sheet_name="Sheet1"):
    """将列表数据写入Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    for row in data:
        ws.append(row)
    
    wb.save(filepath)
    print(f"数据已保存到: {filepath}")

# 测试工具函数
sample_data = [
    ['姓名', '年龄'],
    ['张三', 25],
    ['李四', 30],
]
write_list_to_excel(sample_data, '工具函数示例.xlsx')
read_data = read_excel_to_list('工具函数示例.xlsx')
print("读取的数据:", read_data)

print("\n所有Excel操作示例完成！")